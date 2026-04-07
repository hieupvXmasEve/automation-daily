from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ..models import ComparisonRecord, ExcelOrder, PdfOrder


def export_comparison_excel(
    path: Path,
    comparison_rows: list[ComparisonRecord],
    excel_orders: list[ExcelOrder],
    pdf_orders: list[PdfOrder],
) -> None:
    workbook = Workbook()
    comparison_sheet = workbook.active
    comparison_sheet.title = "comparison"
    _write_sheet(comparison_sheet, [row.to_row() for row in comparison_rows])
    _write_sheet(workbook.create_sheet("excel_orders"), [row.to_row() for row in excel_orders])
    _write_sheet(workbook.create_sheet("pdf_orders"), [row.to_row() for row in pdf_orders])
    workbook.save(path)


def _write_sheet(sheet, rows: list[dict[str, object]]) -> None:
    if not rows:
        sheet.append(["message"])
        sheet.append(["No data"])
        return

    headers = list(rows[0].keys())
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
